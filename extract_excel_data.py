#!/usr/bin/env python3
"""
Extract all data from GAIA_Complete_Tool.xlsx and save as JSON
"""
import json
from openpyxl import load_workbook

def extract_workbook_data(excel_file):
    """Extract all data from Excel workbook"""
    print(f"Loading workbook: {excel_file}")
    wb = load_workbook(excel_file, data_only=True)

    workbook_data = {
        "sheets": [],
        "summary": {
            "total_sheets": 0,
            "sheet_info": []
        }
    }

    for sheet_name in wb.sheetnames:
        print(f"\nProcessing sheet: {sheet_name}")
        ws = wb[sheet_name]

        # Extract all data from the sheet
        sheet_data = []
        max_row = ws.max_row
        max_col = ws.max_column

        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            row_data = []
            for cell in row:
                # Handle different cell types
                value = cell.value
                if value is None:
                    row_data.append("")
                elif isinstance(value, (int, float)):
                    row_data.append(value)
                else:
                    row_data.append(str(value))
            sheet_data.append(row_data)

        # Add sheet to workbook data
        workbook_data["sheets"].append({
            "name": sheet_name,
            "data": sheet_data,
            "rows": len(sheet_data),
            "columns": max_col
        })

        # Update summary
        workbook_data["summary"]["sheet_info"].append({
            "name": sheet_name,
            "rows": len(sheet_data),
            "columns": max_col
        })

        print(f"  ✓ Extracted {len(sheet_data)} rows × {max_col} columns")

    workbook_data["summary"]["total_sheets"] = len(wb.sheetnames)

    return workbook_data

if __name__ == "__main__":
    # Extract data
    data = extract_workbook_data("GAIA_Complete_Tool.xlsx")

    # Save to JSON
    output_file = "src/workbook-data.json"

    # Create src directory if it doesn't exist
    import os
    os.makedirs("src", exist_ok=True)

    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print(f"\n{'='*60}")
    print(f"✓ Data saved to: {output_file}")
    print(f"{'='*60}")
    print(f"\nSUMMARY:")
    print(f"  Total Sheets: {data['summary']['total_sheets']}")
    print(f"\n  Sheet Details:")
    for info in data['summary']['sheet_info']:
        print(f"    • {info['name']}: {info['rows']} rows × {info['columns']} columns")
    print(f"{'='*60}\n")
