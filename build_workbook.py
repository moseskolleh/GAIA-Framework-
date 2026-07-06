#!/usr/bin/env python3
"""
build_workbook.py — deterministic generator of GAIA_Assessment_Tool.xlsx

GAIA 2.0 (FRAMEWORK.md §9): the Excel tool is a BUILD ARTIFACT, generated from
this script plus the CSV data tables in data/. It is never hand-edited.
Re-running this script always produces the same workbook (no runtime
timestamps; the only date is the static VERSION_DATE constant below).

Engine math (FRAMEWORK.md §4, implemented identically in every sheet):
    e_out        = Wh per 1000 output tokens at IT boundary (models.csv;
                   serving-stack multiplier S already included)
    E_IT_request = e_out * (T_out + T_in/10) / 1000                [Wh]
    E_request    = E_IT_request * PUE                              [Wh]
    E_month      = E_request * Q_month / 1000                      [kWh]
    E_IT_month   = E_IT_request * Q_month / 1000                   [kWh]
    C_location   = E_month * CI_location / 1000                    [kg CO2e]
    C_market     = E_month * CI_market / 1000  (only if CI_market given)
    C_embodied   = C_location * embodied_adder (default 0.25)      [kg CO2e]
    Water        = E_IT_month * WUE + E_month * EWIF               [L]
    low/high     = every output scaled by wh_low/e_out and wh_high/e_out
    Grade        = central E_request vs grading.csv bands for the task class
    Frugality    = count of "No" among 3 questions: 0->F0, 1->F1, >=2->F2+

Formulas use INDEX/MATCH (never XLOOKUP) with IFERROR guards, for
compatibility with older Excel and LibreOffice.
"""

import csv
import datetime
import os
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# ----------------------------------------------------------------------------
# Constants
# ----------------------------------------------------------------------------
VERSION = "2.0.0"
VERSION_DATE = "2026-07-06"  # static build/version date — the only date used
LICENSE = "MIT"

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(HERE, "data")
OUT = os.path.join(HERE, "GAIA_Assessment_Tool.xlsx")

# ----------------------------------------------------------------------------
# Styles
# ----------------------------------------------------------------------------
C_HEADER = "1F4E42"      # dark green — section/table headers
C_SUBHEAD = "3D6B5E"     # medium green — sub-headers
C_INPUT = "FFF2CC"       # light yellow — input cells (user edits these)
C_CALC = "DCE6F1"        # light blue — computed cells (do not edit)
C_NOTE = "F2F2F2"        # light gray — notes / documentation

TIER_FILLS = {"T1": "C6EFCE", "T2": "E2EFDA", "T3": "FFF2CC", "T4": "FCE4D6"}
GRADE_STYLE = {  # grade -> (fill, font color)
    "A": ("1E7B34", "FFFFFF"),
    "B": ("70AD47", "FFFFFF"),
    "C": ("FFD966", "000000"),
    "D": ("ED7D31", "FFFFFF"),
    "E": ("C00000", "FFFFFF"),
}

F_TITLE = Font(name="Calibri", size=16, bold=True, color=C_HEADER)
F_SUB = Font(name="Calibri", size=10, italic=True, color="595959")
F_HDR = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
F_BOLD = Font(name="Calibri", size=11, bold=True)
F_BODY = Font(name="Calibri", size=11)
F_SMALL = Font(name="Calibri", size=9, color="595959")

FILL_HDR = PatternFill("solid", fgColor=C_HEADER)
FILL_SUBHEAD = PatternFill("solid", fgColor=C_SUBHEAD)
FILL_INPUT = PatternFill("solid", fgColor=C_INPUT)
FILL_CALC = PatternFill("solid", fgColor=C_CALC)
FILL_NOTE = PatternFill("solid", fgColor=C_NOTE)

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)

NF_WH = "0.0000"     # per-request Wh (4 dp so small models' low bounds stay visible)
NF_KWH = "0.000"     # monthly kWh
NF_KG = "0.000"      # kg CO2e
NF_L = "0.0"         # litres
NF_EQ = "#,##0.0"    # equivalents


# ----------------------------------------------------------------------------
# CSV loading
# ----------------------------------------------------------------------------
def load_csv(name):
    with open(os.path.join(DATA, name), newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


MODELS = load_csv("models.csv")
REGIONS = load_csv("regions.csv")
FACILITIES = load_csv("facilities.csv")
GRADING = load_csv("grading.csv")
MITIGATION = load_csv("mitigation.csv")
EQUIVALENTS = load_csv("equivalents.csv")

N_MODELS = len(MODELS)
N_REGIONS = len(REGIONS)
N_FACILITIES = len(FACILITIES)
N_GRADES = len(GRADING)


def num(s):
    """Parse a CSV numeric field to int/float so Excel stores a number."""
    s = str(s).strip()
    if s == "":
        return ""
    f = float(s)
    return int(f) if f == int(f) and "." not in s and "e" not in s.lower() else f


# ----------------------------------------------------------------------------
# Small helpers
# ----------------------------------------------------------------------------
def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def title_block(ws, title, subtitle):
    ws["A1"] = title
    ws["A1"].font = F_TITLE
    ws["A2"] = subtitle
    ws["A2"].font = F_SUB


def header_row(ws, row, headers, start_col=1):
    for j, h in enumerate(headers, start=start_col):
        c = ws.cell(row=row, column=j, value=h)
        c.font = F_HDR
        c.fill = FILL_HDR
        c.alignment = CENTER_WRAP
        c.border = BORDER


def data_cell(ws, row, col, value, wrap=True, nf=None, fill=None, font=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font or F_BODY
    c.border = BORDER
    if wrap:
        c.alignment = WRAP
    if nf:
        c.number_format = nf
    if fill:
        c.fill = fill
    return c


def section_banner(ws, row, text, span):
    for j in range(2, 2 + span):
        c = ws.cell(row=row, column=j)
        c.fill = FILL_HDR
        c.border = BORDER
    c = ws.cell(row=row, column=2, value=text)
    c.font = F_HDR


# ============================================================================
# ENGINE row registry — single source of truth for all cell addresses
# ============================================================================
# Each entry: (key, label, formula_template_or_value, unit, documentation)
# Formula templates may contain {key} placeholders resolved to $B$<row>.
# Cross-sheet references use eng(key) -> "Engine!$B$<row>".

ENGINE_ROWS = []
ENGINE_ADDR = {}  # key -> row number on Engine sheet


def erow(key, label, formula, unit="", doc=""):
    ENGINE_ROWS.append(("row", key, label, formula, unit, doc))


def esec(text):
    ENGINE_ROWS.append(("sec", None, text, None, None, None))


def eblank():
    ENGINE_ROWS.append(("blank", None, None, None, None, None))


def build_engine_registry():
    # Assessment input cell addresses (must match build_assessment layout)
    esec("1 · INPUTS (echoed from the Assessment sheet)")
    erow("model", "Model", "=IF(Assessment!$C$4=\"\",\"\",Assessment!$C$4)",
         "", "Assessment input 1")
    erow("task_class", "Task class", "=IF(Assessment!$C$5=\"\",\"\",Assessment!$C$5)",
         "", "Assessment input 2")
    erow("t_in", "Tokens in per request (T_in)", "=Assessment!$C$6",
         "tokens", "Assessment input 3")
    erow("t_out", "Tokens out per request (T_out)", "=Assessment!$C$7",
         "tokens", "Assessment input 4")
    erow("q_month", "Queries per month (Q_month)", "=Assessment!$C$8",
         "requests", "Assessment input 5")
    erow("facility", "Facility profile", "=IF(Assessment!$C$9=\"\",\"\",Assessment!$C$9)",
         "", "Assessment input 6")
    erow("region", "Hosting region", "=IF(Assessment!$C$10=\"\",\"\",Assessment!$C$10)",
         "", "Assessment input 7")
    erow("adder", "Embodied adder", "=Assessment!$C$11",
         "fraction", "Assessment input 8 (default 0.25; range 0.10–0.50, §4.6)")
    erow("market_ci", "Market-based CI (optional)",
         "=IF(Assessment!$C$12=\"\",\"\",Assessment!$C$12)",
         "g CO2e/kWh", "Assessment input 9; blank = not provided (P6)")
    erow("q1", "Frugality Q1 — necessity", "=IF(Assessment!$C$14=\"\",\"\",Assessment!$C$14)",
         "", "Yes/No")
    erow("q2", "Frugality Q2 — right-sizing", "=IF(Assessment!$C$15=\"\",\"\",Assessment!$C$15)",
         "", "Yes/No")
    erow("q3", "Frugality Q3 — token discipline", "=IF(Assessment!$C$16=\"\",\"\",Assessment!$C$16)",
         "", "Yes/No")
    eblank()

    m_end = 1 + N_MODELS      # last data row on Model Database
    f_end = 1 + N_FACILITIES
    r_end = 1 + N_REGIONS
    g_end = 1 + N_GRADES

    def mdb(col):
        return (f"=IFERROR(INDEX('Model Database'!${col}$2:${col}${m_end},"
                f"MATCH({{model}},'Model Database'!$A$2:$A${m_end},0)),\"\")")

    def fac(col):
        return (f"=IFERROR(INDEX('Facility Profiles'!${col}$2:${col}${f_end},"
                f"MATCH({{facility}},'Facility Profiles'!$A$2:$A${f_end},0)),\"\")")

    def reg(col):
        return (f"=IFERROR(INDEX('Region Factors'!${col}$2:${col}${r_end},"
                f"MATCH({{region}},'Region Factors'!$A$2:$A${r_end},0)),\"\")")

    def grd(col):
        return (f"=IFERROR(INDEX(Grading!${col}$2:${col}${g_end},"
                f"MATCH({{task_class}},Grading!$A$2:$A${g_end},0)),\"\")")

    esec("2 · LOOKED-UP FACTORS (INDEX/MATCH against the data sheets)")
    erow("e_out", "e_out — model energy, central", mdb("E"),
         "Wh / 1k output tokens",
         "models.csv wh_per_1k_output_tokens_it — IT boundary (accelerator + host); "
         "serving-stack multiplier S already included")
    erow("wh_low", "e_out low bound", mdb("F"), "Wh / 1k output tokens",
         "models.csv wh_low (tier band = default minimum width; some rows are "
         "wider with justification in their basis column)")
    erow("wh_high", "e_out high bound", mdb("G"), "Wh / 1k output tokens",
         "models.csv wh_high (tier band = default minimum width; some rows are "
         "wider with justification in their basis column)")
    erow("tier", "Data-quality tier of model row", mdb("H"), "",
         "T1 measured / T2 disclosed / T3 benchmarked / T4 modeled (§4.5)")
    erow("f_low", "Low scale factor", "=IFERROR({wh_low}/{e_out},\"\")", "×",
         "wh_low / e_out — every output is scaled by this for the low bound")
    erow("f_high", "High scale factor", "=IFERROR({wh_high}/{e_out},\"\")", "×",
         "wh_high / e_out — every output is scaled by this for the high bound")
    erow("pue", "PUE — facility overhead", fac("B"), "×",
         "facilities.csv pue for the chosen profile")
    erow("wue", "WUE — on-site water", fac("E"), "L/kWh (IT energy)",
         "facilities.csv wue_l_kwh for the chosen profile")
    erow("ci", "CI — location-based grid carbon", reg("B"), "g CO2e/kWh",
         "regions.csv ci_location_gco2_kwh for the chosen region")
    erow("ewif", "EWIF — off-site water", reg("E"), "L/kWh (total energy)",
         "regions.csv ewif_l_kwh for the chosen region")
    erow("def_t_in", "Task-class default tokens in", grd("C"), "tokens",
         "grading.csv tokens_in_default — reference for Assessment input 3")
    erow("def_t_out", "Task-class default tokens out", grd("D"), "tokens",
         "grading.csv tokens_out_default — reference for Assessment input 4")
    erow("a_max", "Grade A maximum", grd("E"), "Wh/request",
         "grading.csv grade_a_max_wh for the chosen task class")
    erow("b_max", "Grade B maximum", grd("F"), "Wh/request", "grading.csv grade_b_max_wh")
    erow("c_max", "Grade C maximum", grd("G"), "Wh/request", "grading.csv grade_c_max_wh")
    erow("d_max", "Grade D maximum", grd("H"), "Wh/request",
         "grading.csv grade_d_max_wh; above this = grade E")
    eblank()

    esec("3 · CENTRAL ESTIMATE (§4 equations)")
    erow("e_it_req", "E_IT_request — IT energy per request",
         "=IFERROR({e_out}*({t_out}+{t_in}/10)/1000,\"\")", "Wh",
         "E_IT = e_out × (T_out + T_in/10) / 1000 — input tokens cost ≈1/10 of "
         "output tokens (§4.1, k=10)")
    erow("e_req", "E_request — full energy per request",
         "=IFERROR({e_it_req}*{pue},\"\")", "Wh",
         "E_request = E_IT_request × PUE (§4.1; S is already inside e_out)")
    erow("e_it_month", "E_IT_month — monthly IT energy",
         "=IFERROR({e_it_req}*{q_month}/1000,\"\")", "kWh",
         "E_IT_month = E_IT_request × Q_month / 1000")
    erow("e_month", "E_month — monthly total energy",
         "=IFERROR({e_req}*{q_month}/1000,\"\")", "kWh",
         "E_month = E_request × Q_month / 1000 (§4.2)")
    erow("c_loc", "Carbon, location-based",
         "=IFERROR({e_month}*{ci}/1000,\"\")", "kg CO2e/month",
         "C_op(location) = E_month × CI_location / 1000 (§4.3)")
    erow("c_mkt", "Carbon, market-based",
         "=IF(OR({market_ci}=\"\",NOT(ISNUMBER({market_ci}))),\"not provided\","
         "IFERROR({e_month}*{market_ci}/1000,\"\"))",
         "kg CO2e/month",
         "C_op(market) = E_month × CI_market / 1000 — ONLY if a numeric market CI is "
         "supplied; never blended with location-based (P6)")
    erow("c_emb", "Embodied carbon",
         "=IFERROR({c_loc}*{adder},\"\")", "kg CO2e/month",
         "C_embodied = C_op(location) × embodied_adder (§4.6; default 0.25)")
    erow("water", "Water (two-path)",
         "=IFERROR({e_it_month}*{wue}+{e_month}*{ewif},\"\")", "L/month",
         "W = E_IT_month × WUE + E_month × EWIF (§4.4, Li et al. two-path model)")
    eblank()

    esec("4 · LOW / HIGH BOUNDS (energy-dominated; factor uncertainty adds beyond this)")
    for k, lbl, unit in [
        ("e_req", "Energy per request", "Wh"),
        ("e_month", "Monthly energy", "kWh"),
        ("c_loc", "Carbon, location-based", "kg CO2e/month"),
        ("c_emb", "Embodied carbon", "kg CO2e/month"),
        ("water", "Water", "L/month"),
    ]:
        erow(k + "_low", lbl + " — LOW", "=IFERROR({%s}*{f_low},\"\")" % k, unit,
             "central × (wh_low / e_out)")
        erow(k + "_high", lbl + " — HIGH", "=IFERROR({%s}*{f_high},\"\")" % k, unit,
             "central × (wh_high / e_out)")
    erow("c_mkt_low", "Carbon, market-based — LOW",
         "=IF(OR({market_ci}=\"\",NOT(ISNUMBER({market_ci}))),\"not provided\","
         "IFERROR({c_mkt}*{f_low},\"\"))", "kg CO2e/month",
         "central × (wh_low / e_out); 'not provided' without a numeric market CI")
    erow("c_mkt_high", "Carbon, market-based — HIGH",
         "=IF(OR({market_ci}=\"\",NOT(ISNUMBER({market_ci}))),\"not provided\","
         "IFERROR({c_mkt}*{f_high},\"\"))", "kg CO2e/month",
         "central × (wh_high / e_out); 'not provided' without a numeric market CI")
    eblank()

    esec("5 · GRADE, FRUGALITY FLAG, GUIDANCE (§5)")
    erow("grade", "Efficiency grade",
         "=IF(OR({e_req}=\"\",NOT(ISNUMBER({a_max}))),\"\","
         "IF({e_req}<={a_max},\"A\",IF({e_req}<={b_max},\"B\","
         "IF({e_req}<={c_max},\"C\",IF({e_req}<={d_max},\"D\",\"E\")))))",
         "A–E", "Central E_request vs task-class bands (§5.2); blank unless a "
         "valid task class is selected")
    erow("no_count", "Frugality — count of \"No\" answers",
         "=COUNTIF({q1}:{q3},\"No\")", "0–3", "Across Q1–Q3 (§5.3)")
    erow("flag", "Frugality flag",
         "=IF({no_count}=0,\"F0\",IF({no_count}=1,\"F1\",\"F2+\"))",
         "F0/F1/F2+", "0 No → F0; 1 No → F1; ≥2 No → F2+ — reported BESIDE the "
         "grade, never combined into a composite score")
    erow("guidance", "Guidance",
         "=IF({grade}=\"\",\"\",IF({grade}=\"E\",\"Re-architect: reasoning budgets, "
         "model routing, or task redesign before scale-up\","
         "IF(OR({grade}=\"A\",{grade}=\"B\"),"
         "IF({flag}=\"F0\",\"Proceed; monitor quarterly\","
         "\"Efficient but wasteful by design — fix necessity/right-sizing first\"),"
         "IF({flag}=\"F0\",\"Justified heavy use — apply mitigation levers; set an "
         "intensity-reduction target\",\"Priority for intervention\"))))",
         "", "Grade × flag decision table (§5.4)")
    eblank()

    esec("6 · REAL-WORLD EQUIVALENTS (communication aids — factors from data/equivalents.csv)")
    for i, r in enumerate(EQUIVALENTS):
        erow(f"eqf_{i}",
             f"Factor: {r['quantity']} → {r['equivalent']} (per {r['per_unit']})",
             num(r["factor"]), "", r["source"])
    erow("eq_led", "≈ Hours of a 10 W LED bulb",
         "=IFERROR({e_month}*{eqf_0},\"\")", "h/month",
         "E_month [kWh] × factor")
    erow("eq_phone", "≈ Smartphone full charges",
         "=IFERROR({e_month}*{eqf_2},\"\")", "charges/month",
         "E_month [kWh] × factor")
    erow("eq_car", "≈ Km driven in an average petrol car",
         "=IFERROR({c_loc}*{eqf_3},\"\")", "km/month",
         "C_op(location) [kg] × factor")
    erow("eq_ev", "≈ Km driven in an average EV",
         "=IFERROR({e_month}*{eqf_1},\"\")", "km/month",
         "E_month [kWh] × factor")
    erow("eq_glasses", "≈ Drinking glasses (250 mL)",
         "=IFERROR({water}*{eqf_4},\"\")", "glasses/month",
         "Water [L] × factor")
    erow("eq_shower", "≈ 8-minute showers",
         "=IFERROR({water}/1000*{eqf_5},\"\")", "showers/month",
         "Water [L] / 1000 × factor")

    # Assign row numbers (title=1, note=2, blank=3, content starts at 4)
    r = 4
    for entry in ENGINE_ROWS:
        if entry[0] == "row":
            ENGINE_ADDR[entry[1]] = r
        r += 1


build_engine_registry()


def local(key):
    return f"$B${ENGINE_ADDR[key]}"


def eng(key):
    return f"Engine!$B${ENGINE_ADDR[key]}"


def resolve(template, ref):
    """Fill {key} placeholders using ref(key) -> cell address."""
    out = template
    for k in ENGINE_ADDR:
        out = out.replace("{%s}" % k, ref(k))
    return out


# ============================================================================
# Sheet builders
# ============================================================================
def build_start_here(wb):
    ws = wb.create_sheet("Start Here")
    set_widths(ws, [3, 100])
    ws.sheet_view.showGridLines = False

    ws["B2"] = "GAIA 2.0 — Green AI Assessment Tool"
    ws["B2"].font = Font(name="Calibri", size=20, bold=True, color=C_HEADER)
    ws["B3"] = (f"Version {VERSION} · {VERSION_DATE} · {LICENSE} License · "
                "Generated by build_workbook.py from data/*.csv — never hand-edited "
                "(FRAMEWORK.md §9)")
    ws["B3"].font = F_SUB

    rows = [
        ("", ""),
        ("What this tool is", "sub"),
        ("GAIA answers three questions for an organization that USES AI (via APIs, "
         "hosted deployments, or self-hosted models): (1) how large is the "
         "environmental footprint of a use case — energy, carbon, and water — with "
         "uncertainty bounds and data provenance; (2) is that footprint reasonable "
         "for the task, graded against published measurements; (3) what should "
         "change, ranked by evidence. Method: Ground → Assess → Interpret → Act.", ""),
        ("", ""),
        ("How to use it — 5 steps", "sub"),
        ("1.  Open the Assessment sheet and pick your Model and Task class from the "
         "dropdowns.", ""),
        ("2.  Enter your token profile (tokens in / out per request) and queries per "
         "month — class defaults are shown next to the inputs for reference.", ""),
        ("3.  Choose the Facility profile and Hosting region (keep the honest "
         "defaults — 'Unknown (API default)' and 'Global average' — if you don't "
         "know).", ""),
        ("4.  Answer the three frugality questions (necessity, right-sizing, token "
         "discipline) with Yes/No.", ""),
        ("5.  Read the Results block: low/central/high energy, carbon (dual), and "
         "water; the efficiency grade and frugality flag; and the guidance line. "
         "Then pick levers on the Mitigation sheet. Track months on the Usage Log.", ""),
        ("", ""),
        ("Color key", "sub"),
        ("KEY_INPUT", "key"),
        ("KEY_CALC", "key"),
        ("", ""),
        ("One honest note about uncertainty (Principle P2)", "sub"),
        ("Every result is a low/central/high interval, not a single number. The "
         "interval width comes from the data-quality tier of the chosen model row "
         "(×1.15 measured … ×3 modeled — see Methodology & Sources). The bounds "
         "shown are energy-dominated; uncertainty in PUE, grid carbon intensity, "
         "WUE and EWIF adds beyond them. A point estimate without bounds would be "
         "pseudo-precision.", ""),
        ("", ""),
        ("Where everything lives", "sub"),
        ("Assessment = the one sheet you fill in.  Engine = every intermediate "
         "calculation, documented.  Model Database / Region Factors / Facility "
         "Profiles / Grading / Mitigation = the sourced data tables (from "
         "data/*.csv).  Usage Log = monthly tracking template.  Methodology & "
         "Sources = equations, tiers, boundary, limitations, full source list.  "
         "Framework Comparison = how GAIA maps to SCI, GHG Protocol, AI Energy "
         "Score and others.  Changelog = version history.", ""),
    ]
    r = 5
    for text, kind in rows:
        if kind == "sub":
            c = ws.cell(row=r, column=2, value=text)
            c.font = Font(name="Calibri", size=13, bold=True, color=C_HEADER)
        elif kind == "key":
            if text == "KEY_INPUT":
                c = ws.cell(row=r, column=2,
                            value="  Input cell — light yellow: the cells you edit "
                                  "(dropdowns or typed values)")
                c.fill = FILL_INPUT
            else:
                c = ws.cell(row=r, column=2,
                            value="  Computed cell — light blue: results calculated "
                                  "by the workbook; do not edit")
                c.fill = FILL_CALC
            c.font = F_BODY
            c.border = BORDER
        else:
            c = ws.cell(row=r, column=2, value=text)
            c.font = F_BODY
            c.alignment = WRAP
        r += 1
    return ws


def build_assessment(wb):
    ws = wb.create_sheet("Assessment")
    set_widths(ws, [2, 52, 16, 15, 15, 46])
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws["B1"] = "GAIA 2.0 — AI Use-Case Assessment"
    ws["B1"].font = F_TITLE
    ws["B2"] = ("Edit the yellow cells only. Blue cells are computed from the "
                "Engine sheet. All results are low / central / high (P2).")
    ws["B2"].font = F_SUB

    section_banner(ws, 3, "INPUTS", 5)

    def input_row(r, label, value, note, nf=None):
        data_cell(ws, r, 2, label, font=F_BOLD)
        c = data_cell(ws, r, 3, value, wrap=False, nf=nf, fill=FILL_INPUT)
        c.alignment = CENTER
        n = ws.cell(row=r, column=4, value=note)
        n.font = F_SMALL
        n.alignment = WRAP
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        return c

    g_end = 1 + N_GRADES
    input_row(4, "1 · Model", "GPT-4o",
              "Dropdown — full provenance in 'Model Database'")
    input_row(5, "2 · Task class", "Standard",
              "L / S / H / R — definitions and bands in 'Grading'")
    input_row(6, "3 · Tokens in (per request)", 500, "", nf="0")
    ws["D6"] = (f'=IFERROR("Class default: "&INDEX(Grading!$C$2:$C${g_end},'
                f'MATCH($C$5,Grading!$A$2:$A${g_end},0)),"")')
    ws["D6"].font = F_SMALL
    input_row(7, "4 · Tokens out (per request)", 300, "", nf="0")
    ws["D7"] = (f'=IFERROR("Class default: "&INDEX(Grading!$D$2:$D${g_end},'
                f'MATCH($C$5,Grading!$A$2:$A${g_end},0)),"")')
    ws["D7"].font = F_SMALL
    input_row(8, "5 · Queries per month", 1000, "", nf="#,##0")
    input_row(9, "6 · Facility profile", "Unknown (API default)",
              "PUE and WUE per profile in 'Facility Profiles'")
    input_row(10, "7 · Hosting region", "Global average",
              "Grid CI and EWIF in 'Region Factors'; keep 'Global average' if unknown")
    input_row(11, "8 · Embodied adder", 0.25,
              "Default 0.25 of location-based carbon; sensible range 0.10–0.50 (§4.6)",
              nf="0.00")
    input_row(12, "9 · Market-based CI (g CO2e/kWh) — optional", None,
              "Leave blank unless your provider discloses a market-based factor; "
              "reported beside, never instead of, location-based (P6)", nf="0")

    section_banner(ws, 13, "FRUGALITY CHECK (§5.3) — three auditable Yes/No questions", 5)
    input_row(14, "Q1 · Necessity — does the task need generative AI at all "
                  "(vs. search, template, rules, or a human)?", "Yes", "")
    input_row(15, "Q2 · Right-sizing — has a model ≥1 capability class smaller been "
                  "evaluated on this task, with documented quality results?", "Yes", "")
    input_row(16, "Q3 · Token discipline — are prompts, retrieval payloads and "
                  "reasoning budgets bounded (max-token / thinking caps, caching)?",
              "Yes", "")

    # --- Results ---
    section_banner(ws, 18, "RESULTS — computed on the Engine sheet; do not edit", 5)
    header_row(ws, 19, ["Metric", "Low", "Central", "High", "Unit / note"], start_col=2)

    def res_row(r, label, low, central, high, unit, nf):
        data_cell(ws, r, 2, label, font=F_BOLD)
        for col, v in ((3, low), (4, central), (5, high)):
            c = data_cell(ws, r, col, v, wrap=False, fill=FILL_CALC,
                          nf=nf if v not in ("—",) else None)
            c.alignment = CENTER
        n = data_cell(ws, r, 6, unit, fill=FILL_NOTE)
        n.font = F_SMALL

    res_row(20, "Energy per request",
            f"={eng('e_req_low')}", f"={eng('e_req')}", f"={eng('e_req_high')}",
            "Wh / request (full-stack: IT × PUE)", NF_WH)
    res_row(21, "Energy per month",
            f"={eng('e_month_low')}", f"={eng('e_month')}", f"={eng('e_month_high')}",
            "kWh / month", NF_KWH)
    res_row(22, "Carbon — location-based",
            f"={eng('c_loc_low')}", f"={eng('c_loc')}", f"={eng('c_loc_high')}",
            "kg CO2e / month (physical grid mix)", NF_KG)
    res_row(23, "Carbon — market-based",
            f"={eng('c_mkt_low')}", f"={eng('c_mkt')}", f"={eng('c_mkt_high')}",
            "kg CO2e / month — 'not provided' unless input 9 is set (P6)", NF_KG)
    res_row(24, "Embodied carbon",
            f"={eng('c_emb_low')}", f"={eng('c_emb')}", f"={eng('c_emb_high')}",
            "kg CO2e / month = location-based × adder (§4.6)", NF_KG)
    res_row(25, "Water (on-site + off-site)",
            f"={eng('water_low')}", f"={eng('water')}", f"={eng('water_high')}",
            "L / month (two-path model, §4.4)", NF_L)

    # Grade / flag / tier
    data_cell(ws, 27, 2, "EFFICIENCY GRADE — central Wh/request vs task-class "
                         "bands (§5.2)", font=F_BOLD)
    ws.merge_cells("C27:D28")
    g = ws["C27"]
    g.value = f"={eng('grade')}"
    g.font = Font(name="Calibri", size=28, bold=True)
    g.alignment = CENTER
    for rng in ("C27", "C28", "D27", "D28"):
        ws[rng].border = BORDER
    data_cell(ws, 27, 5, "Frugality flag", fill=FILL_NOTE, font=F_SMALL)
    c = data_cell(ws, 27, 6, f"={eng('flag')}", wrap=False, fill=FILL_CALC,
                  font=Font(name="Calibri", size=14, bold=True))
    c.alignment = CENTER
    data_cell(ws, 28, 5, "Model data tier", fill=FILL_NOTE, font=F_SMALL)
    c = data_cell(ws, 28, 6, f"={eng('tier')}", wrap=False, fill=FILL_CALC,
                  font=F_BOLD)
    c.alignment = CENTER

    data_cell(ws, 30, 2, "GUIDANCE (§5.4 — grade × flag)", font=F_BOLD)
    ws.merge_cells("C30:F30")
    gd = ws["C30"]
    gd.value = f"={eng('guidance')}"
    gd.font = F_BOLD
    gd.alignment = WRAP
    gd.fill = FILL_CALC
    for col in "CDEF":
        ws[f"{col}30"].border = BORDER
    ws.row_dimensions[30].height = 30

    # Grade conditional formatting
    for grade, (fill, color) in GRADE_STYLE.items():
        ws.conditional_formatting.add(
            "C27:D28",
            CellIsRule(operator="equal", formula=[f'"{grade}"'],
                       fill=PatternFill("solid", fgColor=fill),
                       font=Font(size=28, bold=True, color=color)))

    # Equivalents
    section_banner(ws, 32, "REAL-WORLD EQUIVALENTS — communication aids, not "
                           "results (sourced factors on Engine sheet)", 5)
    eq_rows = [
        ("eq_led", "≈ Hours of a 10 W LED bulb", "hours / month"),
        ("eq_phone", "≈ Smartphone full charges", "charges / month"),
        ("eq_car", "≈ Km driven in an average petrol car", "km / month"),
        ("eq_shower", "≈ 8-minute showers", "showers / month"),
    ]
    for i, (key, label, unit) in enumerate(eq_rows):
        r = 33 + i
        data_cell(ws, r, 2, label)
        c = data_cell(ws, r, 3, f"={eng(key)}", wrap=False, fill=FILL_CALC, nf=NF_EQ)
        c.alignment = CENTER
        d = ws.cell(row=r, column=4, value=unit)
        d.font = F_SMALL

    # --- Data validations ---
    m_end = 1 + N_MODELS
    f_end = 1 + N_FACILITIES
    r_end = 1 + N_REGIONS
    dvs = [
        (DataValidation(type="list",
                        formula1=f"='Model Database'!$A$2:$A${m_end}",
                        allow_blank=True, showDropDown=False), "C4"),
        (DataValidation(type="list", formula1=f"=Grading!$A$2:$A${g_end}",
                        allow_blank=True, showDropDown=False), "C5"),
        (DataValidation(type="list",
                        formula1=f"='Facility Profiles'!$A$2:$A${f_end}",
                        allow_blank=True, showDropDown=False), "C9"),
        (DataValidation(type="list",
                        formula1=f"='Region Factors'!$A$2:$A${r_end}",
                        allow_blank=True, showDropDown=False), "C10"),
        (DataValidation(type="list", formula1='"Yes,No"', allow_blank=True,
                        showDropDown=False), "C14:C16"),
        (DataValidation(type="decimal", operator="between", formula1="0.1",
                        formula2="0.5", allow_blank=True,
                        errorTitle="Embodied adder",
                        error="Use a value between 0.10 and 0.50 (§4.6)",
                        showErrorMessage=True), "C11"),
    ]
    for dv, rng in dvs:
        ws.add_data_validation(dv)
        dv.add(rng)
    return ws


def build_engine(wb):
    ws = wb.create_sheet("Engine")
    set_widths(ws, [46, 16, 22, 95])
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False

    ws["A1"] = "GAIA 2.0 Engine — every intermediate calculation, documented"
    ws["A1"].font = F_TITLE
    ws["A2"] = ("All Assessment results reference this sheet. Column D documents each "
                "formula in framework notation (FRAMEWORK.md §4–§5). Generated — do "
                "not edit.")
    ws["A2"].font = F_SUB

    nf_map = {
        "e_out": NF_WH, "wh_low": NF_WH, "wh_high": NF_WH,
        "f_low": "0.000", "f_high": "0.000", "pue": "0.00", "wue": "0.00",
        "ci": "0", "ewif": "0.00",
        "e_it_req": "0.0000", "e_req": "0.0000",
        "e_it_month": NF_KWH, "e_month": NF_KWH,
        "c_loc": NF_KG, "c_emb": NF_KG, "water": NF_L,
        "e_req_low": "0.0000", "e_req_high": "0.0000",
        "e_month_low": NF_KWH, "e_month_high": NF_KWH,
        "c_loc_low": NF_KG, "c_loc_high": NF_KG,
        "water_low": NF_L, "water_high": NF_L,
        "eq_led": NF_EQ, "eq_phone": NF_EQ, "eq_car": NF_EQ, "eq_shower": NF_EQ,
    }

    r = 4
    for entry in ENGINE_ROWS:
        kind = entry[0]
        if kind == "sec":
            for col in range(1, 5):
                ws.cell(row=r, column=col).fill = FILL_SUBHEAD
                ws.cell(row=r, column=col).border = BORDER
            c = ws.cell(row=r, column=1, value=entry[2])
            c.font = F_HDR
        elif kind == "row":
            _, key, label, formula, unit, doc = entry
            data_cell(ws, r, 1, label, font=F_BOLD)
            val = resolve(formula, local) if isinstance(formula, str) else formula
            c = data_cell(ws, r, 2, val, wrap=False, fill=FILL_CALC,
                          nf=nf_map.get(key))
            c.alignment = CENTER
            u = data_cell(ws, r, 3, unit, fill=FILL_NOTE)
            u.font = F_SMALL
            d = data_cell(ws, r, 4, doc, fill=FILL_NOTE)
            d.font = F_SMALL
        r += 1
    return ws


def data_sheet(wb, name, headers, rows, widths, numeric_cols=(), nfs=None,
               tier_col=None):
    """Generic sourced-data sheet from a CSV."""
    ws = wb.create_sheet(name)
    set_widths(ws, widths)
    ws.freeze_panes = "A2"
    header_row(ws, 1, headers)
    for i, row in enumerate(rows):
        r = i + 2
        for j, v in enumerate(row, start=1):
            val = num(v) if j in numeric_cols else v
            nf = (nfs or {}).get(j)
            c = data_cell(ws, r, j, val, nf=nf)
            if j in numeric_cols:
                c.alignment = Alignment(horizontal="center", vertical="top")
            if tier_col and j == tier_col and str(v) in TIER_FILLS:
                c.fill = PatternFill("solid", fgColor=TIER_FILLS[str(v)])
                c.font = F_BOLD
                c.alignment = CENTER
    return ws


def build_model_db(wb):
    headers = ["Model", "Provider", "Capability class", "Reasoning mode",
               "Wh per 1k output tokens (central, IT boundary)", "Low", "High",
               "Tier", "Vintage", "Basis", "Source"]
    rows = [[m["model"], m["provider"], m["capability_class"], m["reasoning_mode"],
             m["wh_per_1k_output_tokens_it"], m["wh_low"], m["wh_high"],
             m["tier"], m["vintage"], m["basis"], m["source"]] for m in MODELS]
    ws = data_sheet(wb, "Model Database", headers, rows,
                    [26, 12, 13, 12, 16, 9, 9, 7, 10, 60, 45],
                    numeric_cols={5, 6, 7}, nfs={5: "0.000", 6: "0.000", 7: "0.000"},
                    tier_col=8)
    note = ws.cell(row=len(rows) + 3, column=1,
                   value="IT boundary = accelerator + host + idle capacity "
                         "(serving-stack multiplier S already included). Energy "
                         "ONLY is stored per model — carbon and water are computed "
                         "from the separate Region/Facility tables (P3). Tier key: "
                         "T1 measured · T2 disclosed · T3 benchmarked · T4 modeled "
                         "(§4.5). Source of truth: data/models.csv.")
    note.font = F_SMALL
    note.alignment = WRAP
    return ws


def build_regions(wb):
    headers = ["Region", "CI location (g CO2e/kWh)", "CI low", "CI high",
               "EWIF (L/kWh)", "EWIF low", "EWIF high", "Vintage", "Source", "Notes"]
    rows = [[r["region"], r["ci_location_gco2_kwh"], r["ci_low"], r["ci_high"],
             r["ewif_l_kwh"], r["ewif_low"], r["ewif_high"], r["vintage"],
             r["source"], r["notes"]] for r in REGIONS]
    ws = data_sheet(wb, "Region Factors", headers, rows,
                    [22, 13, 9, 9, 11, 9, 9, 9, 55, 42],
                    numeric_cols={2, 3, 4, 5, 6, 7},
                    nfs={2: "0", 3: "0", 4: "0", 5: "0.0", 6: "0.0", 7: "0.0"})
    note = ws.cell(row=len(rows) + 3, column=1,
                   value="CI = location-based grid carbon intensity (annual average). "
                         "EWIF = off-site electricity-water intensity factor. Source "
                         "of truth: data/regions.csv.")
    note.font = F_SMALL
    return ws


def build_facilities(wb):
    headers = ["Facility profile", "PUE", "PUE low", "PUE high", "WUE (L/kWh)",
               "WUE low", "WUE high", "Source", "Notes"]
    rows = [[f["profile"], f["pue"], f["pue_low"], f["pue_high"], f["wue_l_kwh"],
             f["wue_low"], f["wue_high"], f["source"], f["notes"]]
            for f in FACILITIES]
    ws = data_sheet(wb, "Facility Profiles", headers, rows,
                    [30, 8, 9, 9, 11, 9, 9, 70, 45],
                    numeric_cols={2, 3, 4, 5, 6, 7},
                    nfs={i: "0.00" for i in range(2, 8)})
    note = ws.cell(row=len(rows) + 3, column=1,
                   value="PUE multiplies IT energy into total facility energy. WUE "
                         "is on-site cooling water per kWh of IT energy. Source of "
                         "truth: data/facilities.csv.")
    note.font = F_SMALL
    return ws


def build_grading(wb):
    ws = wb.create_sheet("Grading")
    set_widths(ws, [18, 7, 13, 13, 10, 10, 10, 10, 10, 60])
    ws.freeze_panes = "A2"
    header_row(ws, 1, ["Task class", "Code", "Default tokens in",
                       "Default tokens out", "A ≤ (Wh)", "B ≤ (Wh)", "C ≤ (Wh)",
                       "D ≤ (Wh)", "E > (Wh)", "Description"])
    for i, g in enumerate(GRADING):
        r = i + 2
        vals = [g["task_class"], g["class_code"], num(g["tokens_in_default"]),
                num(g["tokens_out_default"]), num(g["grade_a_max_wh"]),
                num(g["grade_b_max_wh"]), num(g["grade_c_max_wh"]),
                num(g["grade_d_max_wh"]), num(g["grade_d_max_wh"]),
                g["description"]]
        for j, v in enumerate(vals, start=1):
            c = data_cell(ws, r, j, v)
            if 3 <= j <= 9:
                c.alignment = Alignment(horizontal="center", vertical="top")
        for j, grade in ((5, "A"), (6, "B"), (7, "C"), (8, "D"), (9, "E")):
            fill, color = GRADE_STYLE[grade]
            c = ws.cell(row=r, column=j)
            c.fill = PatternFill("solid", fgColor=fill)
            c.font = Font(bold=True, color=color)

    r = len(GRADING) + 3
    ws.cell(row=r, column=1, value="How the bands work (§5.2)").font = F_BOLD
    expl = [
        "Grades are assigned on CENTRAL-estimate energy per request (full-stack Wh) "
        "against FIXED, LOGARITHMIC (~×3 spacing), TASK-CONDITIONED bands.",
        "Fixed absolute bands — rather than percentile ranks — keep grades stable "
        "over time and auditable; the ~×3 spacing reflects the order-of-magnitude "
        "spreads observed in every benchmark study.",
        "Anchors: provider-disclosed medians (0.24–0.34 Wh/request) anchor grade A/B "
        "for class S; reasoning models measured at >30 Wh/request (Jegham et al.; AI "
        "Energy Score v2: reasoning ≈ 30× standard) anchor grade E.",
        "Requests are graded WITHIN their task class — comparing a reasoning-agent "
        "workload to an autocomplete query is meaningless.",
        "Carbon and water are REPORTED, not graded: they are dominated by region and "
        "facility (P3), which model choice cannot fix.",
        "Source of truth: data/grading.csv.",
    ]
    for i, t in enumerate(expl):
        c = ws.cell(row=r + 1 + i, column=1, value=t)
        c.font = F_SMALL
        c.alignment = WRAP
        ws.merge_cells(start_row=r + 1 + i, start_column=1,
                       end_row=r + 1 + i, end_column=10)
    return ws


def build_mitigation(wb):
    headers = ["Rank", "Lever", "Who controls it", "Measured effect", "Evidence"]
    rows = [[num(m["rank"]), m["lever"], m["who_controls"], m["measured_effect"],
             m["evidence"]] for m in MITIGATION]
    ws = data_sheet(wb, "Mitigation", headers, rows, [6, 50, 14, 60, 55],
                    numeric_cols={1})
    note = ws.cell(row=len(rows) + 3, column=2,
                   value="Evidence-ranked levers: every effect size is a measured "
                         "number with a citation — 'up to' marketing numbers are "
                         "banned (§6). Provider-side levers (PUE, clean-power "
                         "procurement, WUE) belong in vendor management as "
                         "disclosure asks, not user actions. Source of truth: "
                         "data/mitigation.csv.")
    note.font = F_SMALL
    note.alignment = WRAP
    return ws


def build_usage_log(wb):
    ws = wb.create_sheet("Usage Log")
    set_widths(ws, [22, 26, 12, 12, 12, 14, 13, 12, 14, 12])
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Usage Log — monthly tracking template"
    ws["A1"].font = F_TITLE
    ws["A2"] = ("Enter Month, Model, Queries and average tokens; kWh / kg / L are "
                "computed with the same §4 engine math, using the CURRENT Assessment "
                "facility and region selections (PUE, WUE, CI, EWIF from the Engine "
                "sheet). Central estimates only.")
    ws["A2"].font = F_SUB

    header_row(ws, 3, ["Month", "Model", "Queries", "Tokens in avg",
                       "Tokens out avg", "e_out (Wh/1k out)", "Wh / request",
                       "Energy (kWh)", "Carbon loc. (kg CO2e)", "Water (L)"])

    m_end = 1 + N_MODELS

    def formulas(r):
        return [
            f'=IF($B{r}="","",IFERROR(INDEX(\'Model Database\'!$E$2:$E${m_end},'
            f'MATCH($B{r},\'Model Database\'!$A$2:$A${m_end},0)),""))',
            f'=IF(OR($F{r}="",$D{r}="",$E{r}=""),"",'
            f'$F{r}*($E{r}+$D{r}/10)/1000*{eng("pue")})',
            f'=IF(OR($G{r}="",$C{r}=""),"",$G{r}*$C{r}/1000)',
            f'=IF($H{r}="","",$H{r}*{eng("ci")}/1000)',
            f'=IF(OR($H{r}="",$C{r}=""),"",'
            f'($F{r}*($E{r}+$D{r}/10)/1000*$C{r}/1000)*{eng("wue")}'
            f'+$H{r}*{eng("ewif")})',
        ]

    nf_cols = {6: NF_WH, 7: "0.0000", 8: NF_KWH, 9: NF_KG, 10: NF_L}

    # Example row (row 4) — clearly marked, meant to be overwritten
    ex = ["EXAMPLE — overwrite", "GPT-4o", 1000, 500, 300] + formulas(4)
    for j, v in enumerate(ex, start=1):
        fill = FILL_INPUT if j <= 5 else FILL_CALC
        c = data_cell(ws, 4, j, v, wrap=False, fill=fill, nf=nf_cols.get(j))
        c.font = Font(italic=True, color="808080") if j <= 5 else F_BODY
        if j >= 3:
            c.alignment = CENTER

    # 12 clean monthly rows (rows 5–16) — no fabricated data
    for i in range(12):
        r = 5 + i
        for j in range(1, 6):
            c = data_cell(ws, r, j, None, wrap=False, fill=FILL_INPUT)
            c.alignment = CENTER if j >= 3 else Alignment(horizontal="left")
        for j, f in enumerate(formulas(r), start=6):
            c = data_cell(ws, r, j, f, wrap=False, fill=FILL_CALC,
                          nf=nf_cols.get(j))
            c.alignment = CENTER

    # Totals over the 12 clean rows (example row excluded)
    t = 17
    data_cell(ws, t, 1, "TOTAL (12 template rows; example row excluded)",
              font=F_BOLD, wrap=False)
    for j in (8, 9, 10):
        col = get_column_letter(j)
        c = data_cell(ws, t, j, f"=SUM({col}5:{col}16)", wrap=False,
                      fill=FILL_CALC, nf=nf_cols.get(j))
        c.font = F_BOLD
        c.alignment = CENTER

    dv = DataValidation(type="list",
                        formula1=f"='Model Database'!$A$2:$A${m_end}",
                        allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add("B4:B16")
    return ws


def build_methodology(wb):
    ws = wb.create_sheet("Methodology & Sources")
    set_widths(ws, [3, 120])
    ws.sheet_view.showGridLines = False
    ws["B1"] = "Methodology & Sources"
    ws["B1"].font = F_TITLE
    ws["B2"] = ("The §4 equations as implemented in this workbook, the data-quality "
                "tier ladder, the system boundary, limitations, and the full source "
                "list. Authoritative text: FRAMEWORK.md.")
    ws["B2"].font = F_SUB

    r = 4

    def sub(text):
        nonlocal r
        c = ws.cell(row=r, column=2, value=text)
        c.font = Font(name="Calibri", size=13, bold=True, color=C_HEADER)
        r += 1

    def body(text, mono=False):
        nonlocal r
        c = ws.cell(row=r, column=2, value=text)
        c.font = Font(name="Consolas", size=10) if mono else F_BODY
        c.alignment = WRAP
        r += 1

    def gap():
        nonlocal r
        r += 1

    sub("Equations (FRAMEWORK.md §4) — exactly as computed on the Engine sheet")
    body("E_IT_request [Wh]  =  e_out × (T_out + T_in/10) / 1000", mono=True)
    body("    e_out = model energy per 1000 output tokens at the IT boundary "
         "(accelerator + host + idle; the serving-stack multiplier S ≈ 1.7 is "
         "ALREADY INCLUDED in the Model Database values). Input tokens cost "
         "≈ 1/10 of output tokens (k = 10, §4.1).", mono=True)
    body("E_request [Wh]     =  E_IT_request × PUE            (facility overhead, §4.1)",
         mono=True)
    body("E_month [kWh]      =  E_request × Q_month / 1000                     (§4.2)",
         mono=True)
    body("C_op(location)     =  E_month × CI_location / 1000       [kg CO2e]   (§4.3)",
         mono=True)
    body("C_op(market)       =  E_month × CI_market / 1000 — only when a market CI "
         "is supplied; otherwise 'not provided'. Never blended (P6).", mono=True)
    body("C_embodied         =  C_op(location) × embodied_adder   (default 0.25, "
         "range 0.10–0.50, §4.6)", mono=True)
    body("W_month [L]        =  E_IT_month × WUE_site  +  E_month × EWIF_region "
         "  (two-path water, §4.4)", mono=True)
    body("Bounds: low = central × (wh_low/e_out); high = central × (wh_high/e_out). "
         "The interval is energy-dominated; uncertainty in PUE, CI, WUE, EWIF adds "
         "beyond the shown bounds (documented, P2).", mono=True)
    gap()

    sub("Data-quality tiers (§4.5) — the provenance ladder")
    tier_rows = [
        ("Tier", "Basis", "Uncertainty band (central ×/÷)", "Example"),
        ("T1 Measured", "Direct metering of your deployment (CodeCarbon, "
         "nvidia-smi + host power)", "1.15", "Self-hosted Llama with power logging"),
        ("T2 Disclosed", "Provider-published full-stack measurement with stated "
         "boundary", "1.5", "Google Gemini 0.24 Wh/median prompt; Mistral Large 2 LCA"),
        ("T3 Benchmarked", "Independent standardized benchmark (GPU-only, needs S)",
         "2", "AI Energy Score; Jegham et al. API benchmarks"),
        ("T4 Modeled", "Parameter-count physics model (FLOPs → J via hardware "
         "efficiency)", "3", "EcoLogits-style estimate for an undisclosed model"),
    ]
    for i, row in enumerate(tier_rows):
        for j, v in enumerate(row):
            c = ws.cell(row=r, column=2 + j, value=v)
            c.border = BORDER
            c.alignment = WRAP
            if i == 0:
                c.font = F_HDR
                c.fill = FILL_HDR
            else:
                c.font = F_BODY
                tier = row[0].split()[0]
                if j == 0 and tier in TIER_FILLS:
                    c.fill = PatternFill("solid", fgColor=TIER_FILLS[tier])
                    c.font = F_BOLD
        r += 1
    body("The tier band is the default MINIMUM width: low = central ÷ band, "
         "high = central × band; a model row may carry wider bounds where sources "
         "disagree (documented in its basis column) — the stored wh_low/wh_high "
         "are authoritative. Downstream results scale by the energy bounds "
         "(energy-dominated); factor uncertainty (CI, PUE, WUE, EWIF) adds beyond "
         "the displayed band and is disclosed in the factor sheets.")
    gap()

    sub("System boundary (§2)")
    body("In scope: accelerator energy for inference; host CPU/DRAM/storage/idle "
         "(serving-stack multiplier); data-center overhead (PUE); grid carbon "
         "(location- and market-based CI); on-site cooling water + off-site "
         "electricity-generation water; embodied emissions of serving hardware "
         "(amortized adder); amortized training footprint (optional, OFF by default "
         "— it belongs to the provider's inventory under GHG Protocol logic).")
    body("Out of scope (documented per ISO 14044 cut-off rules): end-user devices, "
         "network transmission (typically <5% of a generative-AI request's "
         "footprint), data-center construction, end-of-life — a truncation that "
         "understates the true total by an estimated 5–15%.")
    gap()

    sub("Known limitations (§10, stated per P2)")
    lims = [
        "Closed providers do not disclose per-model serving data; most model rows "
        "are T3/T4 with wide bounds. GAIA makes the opacity itself visible — that "
        "is the point of the tier column.",
        "Benchmarks measure specific hardware/serving configurations; production "
        "fleets differ (batching, quantization, speculative decoding can shift "
        "per-token energy several-fold in either direction).",
        "Annual-average grid CI ignores hourly variation; marginal-emissions "
        "accounting is out of scope for v2.0.",
        "Water data is the weakest link: EWIF varies by basin and season, and water "
        "STRESS context matters as much as volume. GAIA reports volume and flags "
        "stress-region hosting qualitatively.",
        "Rebound effects (efficiency → more usage) are real and unmodeled; the "
        "frugality flag is the partial control.",
    ]
    for t in lims:
        body("•  " + t)
    gap()

    sub("Foundational sources (§11)")
    srcs = [
        ("Measurement & disclosure",
         "Elsworth et al. (Google), 'Measuring the environmental impact of "
         "delivering AI', arXiv:2508.15734 (2025) · Mistral AI × ADEME/Carbone 4, "
         "'LCA of Mistral Large 2' (2025) · OpenAI (Altman), per-query energy "
         "statement (2025) · Epoch AI, 'How much energy does ChatGPT use?' (2025)"),
        ("Benchmarks",
         "Jegham et al., 'How Hungry is AI?', arXiv:2505.09598 (2025) · Luccioni "
         "et al., 'AI Energy Score' v1–v2 (2025) · Luccioni, Jernite & Strubell, "
         "'Power Hungry Processing', FAccT (2024)"),
        ("Water",
         "Li, Yang, Islam & Ren, 'Making AI Less \"Thirsty\"', arXiv:2304.03271 "
         "(2023; CACM 2025) · Macknick et al., operational water factors (2012)"),
        ("Lifecycle & embodied",
         "Luccioni, Viguier & Ligozat, 'BLOOM LCA', arXiv:2211.02001 (2022) · Faiz "
         "et al., 'LLMCarbon', ICLR (2024) · Boavizta server-impact database · "
         "NVIDIA HGX embodied-carbon disclosures"),
        ("Standards",
         "ISO/IEC 21031:2024 (SCI) · GSF 'SCI for AI' (2025) · ISO 14040/14044 · "
         "ITU-T L.1410 · AFNOR SPEC 2314 (2024) · GHG Protocol Scope 2 Guidance · "
         "EU AI Act Art. 40/51/95, Annex XI"),
        ("Context data",
         "Ember 'Global Electricity Review' (2025, 2024 data) · IEA 'Electricity "
         "2025' / 'Energy & AI' (2025) · Uptime Institute 'Global Data Center "
         "Survey' (2025) · operator sustainability reports (Google, Microsoft, "
         "Meta, AWS)"),
    ]
    for cat, txt in srcs:
        c = ws.cell(row=r, column=2, value=cat + ":")
        c.font = F_BOLD
        r += 1
        body(txt)
        gap()
    return ws


def build_comparison(wb):
    headers = ["Framework", "Type", "What it does", "Relation to GAIA"]
    rows = [
        ("SCI — ISO/IEC 21031:2024 (Green Software Foundation)", "Standard (rate)",
         "SCI = (E×I + M)/R per functional unit",
         "GAIA's §4 computes exactly E, I, M, R; any GAIA result restates as an SCI "
         "score (energy → E, location CI → I, embodied → M, functional unit → R)"),
        ("SCI for AI (GSF, ratified 2025)", "Standard extension",
         "SCI adapted to AI lifecycle incl. training/inference split",
         "GAIA Module A is an implementation; GAIA adds water, grading, frugality, "
         "and the decision layer SCI deliberately leaves out"),
        ("AI Energy Score (Hugging Face/Salesforce/Cohere/CMU, 2025–)",
         "Benchmark + rating",
         "Standardized GPU-energy benchmark on H100, 5-star ratings per task",
         "GAIA consumes it as T3 data; GAIA's task-conditioned grades follow its "
         "logic but grade YOUR deployment, not the bare model"),
        ("EcoLogits (GenAI Impact)", "Software library",
         "ISO-14044-based per-request estimates for API models",
         "Peer methodology for GAIA's T4 model; GAIA is spreadsheet-first and adds "
         "the organizational workflow"),
        ("CodeCarbon / Green Algorithms", "Measurement tools",
         "Meter or estimate compute energy/carbon", "GAIA's T1 data source"),
        ("AFNOR SPEC 2314 — Frugal AI (2024)", "Reference framework",
         "Lifecycle methodology + 31 best practices + 'question the need'",
         "GAIA's §5.3 frugality check descends from it; GAIA adds quantitative "
         "grading and uncertainty tiers"),
        ("ITU-T L.1410 / ISO 14040/44", "LCA standards",
         "Boundary and allocation rules for ICT LCA",
         "Govern GAIA's §2 boundary declarations"),
        ("GHG Protocol (Scope 2 guidance, ICT sector)", "Accounting standard",
         "Corporate inventories, dual reporting",
         "GAIA totals feed Scope 2/3 line items; P6 dual reporting is inherited "
         "from it"),
        ("LLMCarbon (ICLR 2024)", "Academic model",
         "End-to-end (training+inference+embodied) carbon prediction",
         "Basis for GAIA's optional §4.7 training attribution"),
        ("EU AI Act (Art. 40/51, Annex XI)", "Regulation",
         "Energy documentation duties for GPAI providers",
         "GAIA's disclosure template (§8) is structured so provider-side answers "
         "slot in when they become available"),
    ]
    ws = data_sheet(wb, "Framework Comparison", headers, list(rows),
                    [42, 20, 50, 70])
    note = ws.cell(row=len(rows) + 3, column=1,
                   value="What none of these provides — and GAIA does — is the "
                         "combination of: (a) uncertainty-tiered estimates usable "
                         "WITHOUT provider cooperation, (b) water alongside carbon, "
                         "(c) task-conditioned grading of deployments, (d) an "
                         "explicit frugality check, and (e) an Excel artifact a "
                         "non-programmer can run. (FRAMEWORK.md §7)")
    note.font = F_SMALL
    note.alignment = WRAP
    ws.merge_cells(start_row=len(rows) + 3, start_column=1,
                   end_row=len(rows) + 3, end_column=4)
    return ws


def build_changelog(wb):
    ws = wb.create_sheet("Changelog")
    set_widths(ws, [12, 12, 110])
    ws.sheet_view.showGridLines = False
    header_row(ws, 1, ["Version", "Date", "Changes"])
    entries = [
        (VERSION, VERSION_DATE,
         "GAIA 2.0 — full science-based rebuild. The v1 composite Environmental "
         "Score (dimensionally incoherent), Benefit Score (subjective weights) and "
         "'Prohibited' decision matrix are REMOVED; replaced by task-conditioned "
         "A–E energy grades on fixed logarithmic bands, a frugality flag (F0/F1/F2+) "
         "reported beside the grade, and grade×flag guidance. Per-model "
         "water/carbon columns abolished (P3): the model database stores energy "
         "only, with source, vintage, tier and bounds; carbon and water are computed "
         "from separate sourced Region/Facility tables. Every result is "
         "low/central/high (P2); dual location/market carbon reporting (P6); "
         "embodied adder (§4.6); two-path water model (§4.4). The workbook is now "
         "GENERATED by build_workbook.py from data/*.csv — never hand-edited (§9). "
         "Full component-by-component audit of what was kept, rebuilt, and removed: "
         "see DECISIONS.md in the repository."),
    ]
    for i, (v, d, txt) in enumerate(entries):
        r = 2 + i
        data_cell(ws, r, 1, v, font=F_BOLD)
        data_cell(ws, r, 2, d)
        data_cell(ws, r, 3, txt)
        ws.row_dimensions[r].height = 150
    note = ws.cell(row=len(entries) + 3, column=3,
                   value="Versioning is semantic (§9): factor-table refreshes bump "
                         "the minor version; methodology changes bump the major "
                         "version and require a documented rationale against P1–P7. "
                         "Errors are fixed in the data tables with a changelog "
                         "entry, never silently.")
    note.font = F_SMALL
    note.alignment = WRAP
    return ws


# ============================================================================
# Main
# ============================================================================
def main():
    wb = Workbook()
    wb.remove(wb.active)

    build_start_here(wb)
    build_assessment(wb)
    build_engine(wb)
    build_model_db(wb)
    build_regions(wb)
    build_facilities(wb)
    build_grading(wb)
    build_mitigation(wb)
    build_usage_log(wb)
    build_methodology(wb)
    build_comparison(wb)
    build_changelog(wb)

    expected = ["Start Here", "Assessment", "Engine", "Model Database",
                "Region Factors", "Facility Profiles", "Grading", "Mitigation",
                "Usage Log", "Methodology & Sources", "Framework Comparison",
                "Changelog"]
    assert wb.sheetnames == expected, wb.sheetnames

    wb.properties.creator = "GAIA build_workbook.py"
    wb.properties.title = "GAIA 2.0 Assessment Tool"
    wb.properties.description = (f"Generated from data/*.csv by build_workbook.py "
                                 f"— version {VERSION} ({VERSION_DATE})")
    # Deterministic metadata: fixed document timestamps from VERSION_DATE.
    fixed = datetime.datetime(*[int(x) for x in VERSION_DATE.split("-")], 0, 0, 0)
    wb.properties.created = fixed
    wb.properties.modified = fixed

    wb.save(OUT)
    normalize_zip(OUT, fixed)
    print(f"Wrote {OUT}")
    print(f"Sheets: {', '.join(expected)}")


def normalize_zip(path, fixed_dt):
    """Rewrite the xlsx zip with fixed entry timestamps so that re-running the
    build produces a byte-identical file (openpyxl stamps entries with the
    current time)."""
    dt = (fixed_dt.year, fixed_dt.month, fixed_dt.day, 0, 0, 0)
    tmp = path + ".tmp"
    with zipfile.ZipFile(path, "r") as zin, \
            zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            info = zipfile.ZipInfo(item.filename, date_time=dt)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = item.external_attr
            zout.writestr(info, zin.read(item.filename))
    os.replace(tmp, path)


if __name__ == "__main__":
    main()
